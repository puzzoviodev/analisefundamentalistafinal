from openpyxl import Workbook
from openpyxl.styles import Font

# Função de classificação com cores
def classificar_divida(valor):
    if valor == 0:
        return "Ótimo", Font(color="00B050")  # Verde escuro
    elif 0 < valor <= 1:
        return "Muito Bom", Font(color="00C000")  # Verde claro
    elif 1 < valor <= 2:
        return "Bom", Font(color="92D050")  # Verde amarelado
    elif 2 < valor <= 3:
        return "Moderado", Font(color="FFFF00")  # Amarelo
    elif 3 < valor <= 4:
        return "Ruim", Font(color="FFC000")  # Laranja
    elif 4 < valor <= 5:
        return "Crítico", Font(color="FF0000")  # Vermelho
    elif valor > 5:
        return "Muito Crítico", Font(color="C00000")  # Vermelho escuro
    else:
        return "Erro", Font(color="000000")  # Preto

# Criar planilha
wb = Workbook()
ws = wb.active
ws.title = "Classificação"

# Valor de teste
valor_teste = 5.8
classificacao, fonte = classificar_divida(valor_teste)

# Escrever na célula com fonte colorida
ws["A1"] = f"Valor: {valor_teste}"
ws["B1"] = f"Classificação: {classificacao}"
ws["B1"].font = fonte

# Salvar arquivo
wb.save("classificacao_divida.xlsx")
